import openpyxl
from flask import Flask, request
import datetime
import json
import os.path

def make_book(sheet):
	sheet.cell(row = 1,column = 1).value = 'N'
	sheet.cell(row = 1,column = 2).value = 'User ID'
	sheet.cell(row = 1,column = 3).value = 'Datetime'
	sheet.cell(row = 1,column = 4).value = 'Item'
	sheet.cell(row = 1,column = 5).value = 'Price'
	return sheet



def make_excel_file(): #тут записываем в наш эксель покупки и сохраняем в файл
	global row_to_write_sells
	global remember_our_sells
	book = openpyxl.load_workbook('data.xlsx')
	sheet = book.active
	for sells in remember_our_sells:
		for i in range(1,6):
			sheet.cell(row_to_write_sells,i).value = sells[i]
		row_to_write_sells+=1
	for i in range (len(remember_our_sells)):
			remember_our_sells.pop(i)
	book.save('data.xlsx')
	book.close


app = Flask(__name__)  


@app.route('/', methods = ['POST', 'GET']) #декоратор
def index():		#обработчик
	our_json= request.get_json()
	global N
	global remember_our_sells
	now_time = datetime.datetime.now().time()
	for sells in our_json['check']:
		temp.append(N)
		temp.append(our_json['User_id'])
		temp.append(now_time)
		temp.append(sells['Item'])
		temp.append(sells['Price'])
		remember_our_sells.append(temp)
		N += 1
		N = str(N)
	if len(remember_our_sells)>1000:
		make_excel_file()
	return 'OK'


if __name__ == "__main__":
	global N
	global remember_our_sells
	global row_to_write_sells
	row_to_write_sells = 2
	remember_our_sells = []
	N = str(1)
	check_file = os.path.exists('data.xlsx')
	if (check_file == False):  
		book = openpyxl.Workbook()
		sheet = book.active
		sheet = make_book(sheet)
		book.save('data.xlsx')
		book.close

	book = openpyxl.load_workbook('data.xlsx')
	sheet = book.active
	
	now_time = datetime.datetime.now().time()


	book.save('data.xlsx')
	book.close
	app.run()
